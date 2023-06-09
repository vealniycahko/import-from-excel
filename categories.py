from openpyxl import load_workbook

from config import get_pg_connection, SPREADSHEET_PATH


wb = load_workbook(SPREADSHEET_PATH)
categories = wb['Категории']

with get_pg_connection() as conn, conn.cursor() as cur:
    cur.execute(""" CREATE EXTENSION IF NOT EXISTS "uuid-ossp";
                
                DROP TABLE IF EXISTS categories CASCADE;
                
                CREATE TABLE categories (
                    id UUID DEFAULT uuid_generate_v4() PRIMARY KEY,
                    title TEXT
                ); """)

for row in categories.iter_rows(values_only=True):
    if row[0]:    
        query = """ INSERT INTO categories (title) VALUES (%s); """
        with get_pg_connection() as conn, conn.cursor() as cur:
            cur.execute(query, (row[0],))
    else:
        continue
