from openpyxl import load_workbook
import json
import re

from config import get_pg_connection, SPREADSHEET_PATH


wb = load_workbook(SPREADSHEET_PATH)
questions = wb['Вопросы']

with get_pg_connection() as conn, conn.cursor() as cur:
    cur.execute(""" CREATE EXTENSION IF NOT EXISTS "uuid-ossp";
                
                DROP TABLE IF EXISTS questions CASCADE;
                
                CREATE TABLE IF NOT EXISTS questions (
                    id UUID DEFAULT uuid_generate_v4() PRIMARY KEY,
                    skill UUID REFERENCES skills(id),
                    complexity SMALLINT,
                    one_option BOOLEAN,
                    question TEXT,
                    variants JSONB,
                    answer SMALLINT[]
                ); """)

for row in questions.iter_rows(min_row=2, values_only=True):
    question = row[3]
    if question:
        with get_pg_connection() as conn, conn.cursor() as cur:
            query = """ SELECT id FROM skills WHERE title = %s; """
            cur.execute(query, (row[0],))
            result = cur.fetchone()
            
        if result:
            skill_id = result[0]
        else:
            print('Нет такого навыка: ', row[0], ' | ', question)
            continue
        
        if row[1] == 'one':
            one_option = True
        elif row[1] == 'many':
            one_option = False
        else:
            print('Нет количества вариантов: ', ' | ', question)
            continue        
    
        if row[2]:
            complexity = int(row[2])
        else:
            print('Сложность не определена', ' | ', question)
            continue
            
        v1 = row[4]
        v2 = row[5]
        v3 = row[6]
        v4 = row[7]
        v5 = row[8]
        v6 = row[9]
        
        variants = {}
        
        if v1:
            variants['1'] = v1
        if v2:
            variants['2'] = v2
        if v3:
            variants['3'] = v3
        if v4:
            variants['4'] = v4
        if v5:
            variants['5'] = v5
        if v6:
            variants['6'] = v6
            
        variants_json = json.dumps(variants)
        
        if row[10]:
            answer = re.findall(r'\d+', row[10])
            answer = [int(n) for n in answer]
        else:
            print('Нет правильного ответа', ' | ', question)
            continue
                
        with get_pg_connection() as conn, conn.cursor() as cur:
            query = """ INSERT INTO questions (skill, complexity, one_option, question, variants, answer)
                VALUES (%s, %s, %s, %s, %s, %s); """
            cur.execute(query, (skill_id, complexity, one_option, question, variants_json, answer))
            
    else:
        continue    
