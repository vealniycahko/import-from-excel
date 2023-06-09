from openpyxl import load_workbook

from config import get_pg_connection, SPREADSHEET_PATH


wb = load_workbook(SPREADSHEET_PATH)
roles = wb['Роли']

with get_pg_connection() as conn, conn.cursor() as cur:
    cur.execute(""" CREATE EXTENSION IF NOT EXISTS "uuid-ossp";
                
                DROP TABLE IF EXISTS roles CASCADE;
                DROP TABLE IF EXISTS roles_to_skills CASCADE;
                
                CREATE TABLE roles (
                    id UUID DEFAULT uuid_generate_v4() PRIMARY KEY,
                    role TEXT,
                    specialization TEXT
                ); 
                
                CREATE TABLE roles_to_skills (
                    role_id UUID REFERENCES roles(id),
                    skill_id UUID REFERENCES skills(id)
                ); """)

for row in roles.iter_rows(min_row=2, values_only=True):
    role = row[0]
    if role:       
        specialization = row[1]
        if not specialization:
            print('Нет специализации: | ', role)
            continue
      
        
        with get_pg_connection() as conn, conn.cursor() as cur:
            query = """ INSERT INTO roles (role, specialization)
                VALUES (%s, %s)
                RETURNING id; """
            cur.execute(query, (role, specialization))
            role_id = cur.fetchone()[0]
        
        
        skills = row[2]
        if skills:
            skills_list = skills.splitlines()
        else:
            print('Нет списка навыков: | ', role, ' - ', specialization)
            continue
        with get_pg_connection() as conn, conn.cursor() as cur:
            select_query = """ SELECT id FROM skills WHERE title = %s; """
            insert_query = """ INSERT INTO roles_to_skills (role_id, skill_id)
                    VALUES (%s, %s); """
            for skill in skills_list:
                cur.execute(select_query, (skill,))
                result = cur.fetchone()
                if result:
                    skill_id = result[0]
                else:
                    print('Нет такого навыка: ', skill, ' | ', role, ' - ', specialization)
                    continue
                cur.execute(insert_query, (role_id, skill_id))
    else:
        continue    
