from openpyxl import load_workbook

from config import get_pg_connection, SPREADSHEET_PATH


wb = load_workbook(SPREADSHEET_PATH)
courses = wb['Курсы']

with get_pg_connection() as conn, conn.cursor() as cur:
    cur.execute(""" CREATE EXTENSION IF NOT EXISTS "uuid-ossp";
                
                DROP TABLE IF EXISTS courses CASCADE;
                DROP TABLE IF EXISTS courses_to_skills CASCADE;
                
                CREATE TABLE courses (
                    id UUID DEFAULT uuid_generate_v4() PRIMARY KEY,
                    title TEXT,
                    content TEXT,
                    link TEXT,
                    hours SMALLINT,
                    category UUID REFERENCES categories(id)
                ); 
                
                CREATE TABLE courses_to_skills (
                    course_id UUID REFERENCES courses(id),
                    skill_id UUID REFERENCES skills(id)
                ); """)

for row in courses.iter_rows(min_row=2, values_only=True):
    title = row[1]
    if title:       
        content = row[2]
        if not content:
            print('Нет содержания курса: | ', title)
            continue
        
        link = row[3]
        # if not link:
        #     print('Нет ссылки на курс: | ', title)
        #     continue
        
        hours = int(row[4])
        if not hours:
            print('Нет длительности курса: | ', title)
            continue
        
        with get_pg_connection() as conn, conn.cursor() as cur:
            query = """ SELECT id FROM categories WHERE title = %s; """
            cur.execute(query, (row[5],))
            result = cur.fetchone()
        if result:
            category = result[0]
        else:
            print('Нет такой категории: ', row[5], ' | ', title)
            continue
        
        
        with get_pg_connection() as conn, conn.cursor() as cur:
            query = """ INSERT INTO courses (title, content, link, hours, category)
                VALUES (%s, %s, %s, %s, %s)
                RETURNING id; """
            cur.execute(query, (title, content, link, hours, category))
            course_id = cur.fetchone()[0]
        
        
        skills = row[6]
        if skills:
            skills_list = skills.splitlines()
        else:
            print('Нет списка навыков: | ', title)
            continue
        with get_pg_connection() as conn, conn.cursor() as cur:
            select_query = """ SELECT id FROM skills WHERE title = %s; """
            insert_query = """ INSERT INTO courses_to_skills (course_id, skill_id)
                    VALUES (%s, %s); """
            for skill in skills_list:
                cur.execute(select_query, (skill,))
                result = cur.fetchone()
                if result:
                    skill_id = result[0]
                else:
                    print('Нет такого навыка: ', skill, ' | ', title)
                    continue
                cur.execute(insert_query, (course_id, skill_id))
    else:
        continue    
